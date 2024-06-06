import openpyxl
import sys
import os

def compare_excel_files(new_file_path, old_file_path):
    wb_new = openpyxl.load_workbook(new_file_path)
    wb_old = openpyxl.load_workbook(old_file_path)
    
    sheet_new = wb_new.active
    sheet_old = wb_old.active

    changes = []

    # Read dates from row 5, columns E to K
    dates = [sheet_new.cell(row=5, column=col).value.date() for col in range(5, 12)]

    for row in range(7, 45):  # Rows 7 to 44
        for col in range(5, 12):  # Columns E to K
            new_value = sheet_new.cell(row=row, column=col).value
            old_value = sheet_old.cell(row=row, column=col).value

            if new_value != old_value:
                # Only include changes to regular cell values
                if not isinstance(old_value, openpyxl.worksheet.formula.ArrayFormula) and not isinstance(new_value, openpyxl.worksheet.formula.ArrayFormula):
                    # Extract the relevant date
                    date = dates[col - 5]

                    # Extract start and end times
                    start_time = sheet_new.cell(row=row, column=2).value
                    end_time = sheet_new.cell(row=row, column=4).value

                    changes.append({
                        "date": date,
                        "start_time": start_time,
                        "end_time": end_time,
                        "old_value": old_value,
                        "new_value": new_value
                    })

    with open('changes.txt', 'w') as f:
        for change in changes:
            f.write(f"{change['date']}, {change['start_time']} to {change['end_time']} changed from '{change['old_value']}' to '{change['new_value']}'\n")
    
    if changes:
        with open(os.environ['GITHUB_ENV'], 'a') as env_file:
            env_file.write("changes_detected=true\n")
    else:
        with open(os.environ['GITHUB_ENV'], 'a') as env_file:
            env_file.write("changes_detected=false\n")

if __name__ == "__main__":
    new_file_path = sys.argv[1]
    old_file_path = sys.argv[2]
    
    compare_excel_files(new_file_path, old_file_path)
